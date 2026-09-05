from pathlib import Path

from openpyxl import load_workbook

src = Path(r"c:\Users\Admin\Desktop\approval\PO_approvalapp\public\june 26 Group EBIDTA.xlsx")
out = Path(r"c:\Users\Admin\Desktop\approval\PO_approvalapp\POApprovalAPI\Scripts\LoadJuneGroupEbitda.sql")

stock_map = {
    "PIL-1": "Plastene India Limited",
    "PIL-2": "Plastene India Limited (Unit -II)",
    "HPBL4": "HCP Plastene Bulkpack Ltd (Unit - IV)",
    "KPW-1": "K.P. WOVEN PRIVATE LIMITED",
    "KPW-2": "K.P. WOVEN PRIVATE LIMITED (UNIT-II)",
    "KPW-3": "K.P. WOVEN PRIVATE LIMITED (UNIT-III)",
    "HPBL-1": "HCP Plastene Bulkpack Ltd",
    "HPBL-2": "HCP Plastene Bulkpack Ltd (Unit - II)",
    "HPBL-3": "HCP Plastene Bulkpack Ltd (Unit - III)",
    "HPBL VAD": "HCP Plastene Bulkpack ltd (Vadodara)",
    "PPL": "Plastene Polyfilms Limited",
    "OEL-1": "Oswal Extrusion Limited",
}
cat_map = {
    "Raw Materials": "RawMaterial",
    "Packing Materials": "Packing",
    "WIP": "WIP",
    "FG": "FinishedGoods",
    "Stores & Consumables": "Stores",
}


def esc(s: str) -> str:
    return (s or "").replace("'", "''").strip()


def entry_type(ledger: str) -> str:
    t = ledger.lower()
    if "expense" in t:
        return "Expense"
    if "income" in t:
        return "Income"
    return "Expense"


wb = load_workbook(src, data_only=True)
ws = wb["Stock sheet"]
stock_rows = []
cur = None
for r in range(1, 97):
    a = ws.cell(r, 1).value
    if a in cat_map:
        cur = cat_map[a]
        continue
    if cur and a in stock_map:
        co = stock_map[a]
        for col, dt in ((2, "2026-03-01"), (3, "2026-04-01"), (4, "2026-05-01"), (5, "2026-06-01")):
            v = ws.cell(r, col).value
            amt = float(v) if isinstance(v, (int, float)) else 0.0
            stock_rows.append((co, dt, cur, amt))

for co in stock_map.values():
    for dt in ("2026-03-01", "2026-04-01", "2026-05-01", "2026-06-01"):
        stock_rows.append((co, dt, "Traded", 0.0))

lines = []
a = lines.append
a("/* Generated from public/june 26 Group EBIDTA.xlsx")
a("   Stock sheet: op -> 2026-03-01, then Apr/May/Jun closings.")
a("   Provisions: 04/05/06.conso TB section A='Provisions'. Amounts in rupees.")
a("   Review CompanyName against FactoryInfo before COMMIT.")
a("   Does not load July+.")
a("*/")
a("SET NOCOUNT ON;")
a("SET XACT_ABORT ON;")
a("BEGIN TRAN;")
a("")
a("IF OBJECT_ID(N'dbo.PnlStockValue', N'U') IS NULL")
a("BEGIN")
a("    RAISERROR(N'PnlStockValue is missing.', 16, 1);")
a("    RETURN;")
a("END")
a("")

for co, dt, cat, amt in stock_rows:
    a(
        "IF EXISTS (SELECT 1 FROM dbo.PnlStockValue "
        f"WHERE CompanyName = N'{esc(co)}' AND StockMonth = '{dt}' "
        f"AND Category = N'{cat}' AND PlantName IS NULL)"
    )
    a("    UPDATE dbo.PnlStockValue")
    a(f"    SET AmountLacs = {amt:.6f}, UploadedAt = GETDATE(), Remarks = N'June Group EBITDA Stock sheet'")
    a(f"    WHERE CompanyName = N'{esc(co)}' AND StockMonth = '{dt}' AND Category = N'{cat}' AND PlantName IS NULL;")
    a("ELSE")
    a("    INSERT INTO dbo.PnlStockValue (CompanyName, PlantName, StockMonth, Category, AmountLacs, Remarks, UploadedAt)")
    a(
        f"    VALUES (N'{esc(co)}', NULL, '{dt}', N'{cat}', {amt:.6f}, "
        f"N'June Group EBITDA Stock sheet', GETDATE());"
    )
    a("")

sheets = [
    ("04.conso TB", "2026-04-01"),
    ("05.conso TB", "2026-05-01"),
    ("06.conso TB", "2026-06-01"),
]
for sheet, dt in sheets:
    ws = wb[sheet]
    comps = []
    for c in range(7, 30):
        v = ws.cell(1, c).value
        if v:
            comps.append((c, str(v).strip()))
    start = None
    for r in range(1, 280):
        if str(ws.cell(r, 1).value or "").strip() == "Provisions":
            start = r + 1
            break
    if start is None:
        raise SystemExit(f"No Provisions header on {sheet}")
    last = ws.max_row or start
    by_co = {co: [] for _, co in comps}
    for r in range(start, last + 1):
        led = ws.cell(r, 2).value
        if not led or not str(led).strip():
            continue
        led = str(led).strip()
        for c, co in comps:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and abs(v) > 0.5:
                by_co[co].append((led, float(v)))
    a(f"-- {sheet} -> {dt}")
    for _, co in comps:
        items = by_co[co]
        if not items:
            continue
        a(
            f"DELETE FROM dbo.Provisioning WHERE LTRIM(RTRIM(companyname)) = N'{esc(co)}' "
            f"AND sysdate >= '{dt}' AND sysdate < DATEADD(month, 1, '{dt}');"
        )
        for led, amt in items:
            a(
                "INSERT INTO dbo.Provisioning (companyname, sysdate, Ledgername, Amount, remarks, EntryType) "
                f"VALUES (N'{esc(co)}', '{dt}', N'{esc(led)[:100]}', {amt:.4f}, "
                f"N'June Group EBITDA {sheet}', N'{entry_type(led)}');"
            )
        a("")

a("COMMIT;")
a("-- ROLLBACK;")
wb.close()
out.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"wrote {out} lines={len(lines)} bytes={out.stat().st_size}")
