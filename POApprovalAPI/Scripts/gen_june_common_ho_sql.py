from pathlib import Path

from openpyxl import load_workbook

src = Path(r"c:\Users\Admin\Desktop\approval\PO_approvalapp\public\june 26 Group EBIDTA.xlsx")
out = Path(r"c:\Users\Admin\Desktop\approval\PO_approvalapp\POApprovalAPI\Scripts\LoadJuneGroupCommonHo.sql")

# Month column on Apr-26 / May-26 / Jun-26 (not YTD). Amounts already in lacs.
unit_map = {
    "PIL I": "Plastene India Limited",
    "PIL II": "Plastene India Limited (Unit -II)",
    "HPBL4": "HCP Plastene Bulkpack Ltd (Unit - IV)",
    "OSWAL - KASEZ": "Oswal Extrusion Limited",
    "PLASTENE POLYFILM LTD.": "Plastene Polyfilms Limited",
    "K P WOVEN PVT LTD - UNIT - I": "K.P. WOVEN PRIVATE LIMITED",
    "K P WOVEN PVT LTD - UNIT -III": "K.P. WOVEN PRIVATE LIMITED (UNIT-III)",
    "HCP PLASTENE BULKPACK LTD. - UNIT - I": "HCP Plastene Bulkpack Ltd",
    "HCP PLASTENE BULKPACK LTD. - UNIT - II": "HCP Plastene Bulkpack Ltd (Unit - II)",
    "HCP PLASTENE BULKPACK LTD. - UNIT - III": "HCP Plastene Bulkpack Ltd (Unit - III)",
}

# P&L also has PIL III / PIL HO / HCP / OCPL — only load if we have a factory name.
extra_map = {
    "PIL HO": "Plastene India Limited - HO",
}

sheets = [
    ("Apr-26", "2026-04-01"),
    ("May-26", "2026-05-01"),
    ("Jun-26", "2026-06-01"),
]


def num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def esc(s: str) -> str:
    return (s or "").replace("'", "''").strip()


wb = load_workbook(src, data_only=True)
ws0 = wb["Jun-26"]
col_by_unit = {}
for c in range(1, 90):
    name = ws0.cell(21, c).value
    if name:
        col_by_unit[str(name).strip()] = c

rows = []  # company, month, common, ho
unknown = []
for sheet, dt in sheets:
    ws = wb[sheet]
    for excel_name, erp in {**unit_map, **extra_map}.items():
        c = col_by_unit.get(excel_name)
        if not c:
            unknown.append(excel_name)
            continue
        common = num(ws.cell(78, c).value)
        ho = num(ws.cell(79, c).value)
        rows.append((erp, dt, excel_name, common, ho, sheet))

print("unknown units", sorted(set(unknown)))
print("---- extracted ----")
for r in rows:
    print(f"{r[1]}  {r[2]:40} common={r[3]:10.4f}  ho={r[4]:10.4f}")

lines = []
a = lines.append
a("/* Generated from public/june 26 Group EBIDTA.xlsx")
a("   Apr-26 / May-26 / Jun-26 rows 78-79 (Common Expenses, HO Expenses), MONTH column.")
a("   Amounts are already in lacs. Upserts dbo.PnlOverhead.")
a("   June Common/HO on Jun-26 is blank (0) for every unit — month June EBITDA will not move.")
a("   April and May allocations are loaded and affect YTD only when viewing Jun 2026.")
a("   Does not load July+ or consolidation columns.")
a("*/")
a("SET NOCOUNT ON;")
a("SET XACT_ABORT ON;")
a("BEGIN TRAN;")
a("")
a("IF OBJECT_ID(N'dbo.PnlOverhead', N'U') IS NULL")
a("BEGIN")
a("    CREATE TABLE dbo.PnlOverhead (")
a("        Id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,")
a("        CompanyName VARCHAR(150) NOT NULL,")
a("        OverheadMonth DATE NOT NULL,")
a("        CommonLacs FLOAT NOT NULL CONSTRAINT DF_PnlOverhead_Common DEFAULT (0),")
a("        HoLacs FLOAT NOT NULL CONSTRAINT DF_PnlOverhead_Ho DEFAULT (0),")
a("        UploadedBy VARCHAR(100) NULL,")
a("        UploadedAt DATETIME NOT NULL CONSTRAINT DF_PnlOverhead_Uploaded DEFAULT (GETDATE()),")
a("        Remarks VARCHAR(500) NULL,")
a("        CONSTRAINT UQ_PnlOverhead UNIQUE (CompanyName, OverheadMonth)")
a("    );")
a("END")
a("")

for erp, dt, excel_name, common, ho, sheet in rows:
    a(f"-- {sheet} {excel_name}")
    a(
        "IF EXISTS (SELECT 1 FROM dbo.PnlOverhead "
        f"WHERE LTRIM(RTRIM(CompanyName)) = N'{esc(erp)}' AND OverheadMonth = '{dt}')"
    )
    a("    UPDATE dbo.PnlOverhead")
    a(f"    SET CommonLacs = {common:.6f}, HoLacs = {ho:.6f}, UploadedAt = GETDATE()")
    a(f"    WHERE LTRIM(RTRIM(CompanyName)) = N'{esc(erp)}' AND OverheadMonth = '{dt}';")
    a("ELSE")
    a("    INSERT INTO dbo.PnlOverhead (CompanyName, OverheadMonth, CommonLacs, HoLacs, UploadedAt)")
    a(f"    VALUES (N'{esc(erp)}', '{dt}', {common:.6f}, {ho:.6f}, GETDATE());")
    a("")

a("COMMIT;")
a("-- ROLLBACK;")
wb.close()
out.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"wrote {out} lines={len(lines)}")
